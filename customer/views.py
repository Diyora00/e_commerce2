from django.views.generic import TemplateView
from django.views import View
import openpyxl
import csv
import json
from django.http import HttpResponse
from django.shortcuts import render, redirect
from customer.models import Customer
from customer.forms import *
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


""" Class based views """


class CustomerListView(View):
    def get(self, request):
        search_post = request.GET.get('search')
        if search_post:
            customers = Customer.objects.filter(Q(fullname__icontains=search_post) | Q(email__icontains=search_post))
        else:
            customers = Customer.objects.all().order_by('-id')

        paginator = Paginator(customers, 2)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.get_page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        context = {'customers': customers, 'page_obj': page_obj}
        return render(request, 'customer/customers.html', context)


class CustomerDetailTemplateView(TemplateView):
    template_name = 'customer/customer-details.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = Customer.objects.get(pk=self.kwargs['customer_id'])
        context['customer'] = customer
        return context


class AddCustomerTemplateView(TemplateView):
    template_name = 'customer/add-customer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = CustomerModelForm()
        return context

    def post(self, request,  *args, **kwargs):
        context = super().get_context_data(**kwargs)

        form = CustomerModelForm(request.POST, request.FILES)
        context['form'] = form
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, 'Customer successfully added!.')
            return redirect('customers')


class UpdateCustomerTemplateView(TemplateView):
    template_name = 'customer/update-customer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = Customer.objects.get(id=kwargs['customer_id'])
        context['form'] = CustomerModelForm(instance=product)
        return context

    def post(self, request,  *args, **kwargs):
        context = super().get_context_data(**kwargs)
        product = Customer.objects.get(id=kwargs['customer_id'])

        form = CustomerModelForm(instance=product, data=request.POST, files=request.FILES)
        context['form'] = form
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, 'Customer successfully added!.')
            return redirect('customers')


class DeleteCustomerTemplateView(TemplateView):
    template_name = 'customer/delete-customer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = Customer.objects.get(id=kwargs['customer_id'])
        context['customer'] = customer
        return context

    def post(self, request, *args, **kwargs):
        customer = Customer.objects.get(id=kwargs['customer_id'])
        customer.delete()
        messages.add_message(request, messages.SUCCESS, 'Customer successfully deleted!')
        return redirect('customers')


""" Function based views """


def show_customers(request):
    search_post = request.GET.get('search')
    if search_post:
        customers = Customer.objects.filter(Q(fullname__icontains=search_post) | Q(email__icontains=search_post))
    else:
        customers = Customer.objects.all().order_by('-id')

    paginator = Paginator(customers, 2)
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {'customers': customers, 'page_obj': page_obj}
    return render(request, 'customer/customers.html', context)


def customer_details(request, customer_id):
    customer = Customer.objects.get(pk=customer_id)
    context = {'customer': customer}
    return render(request, 'customer/customer-details.html', context)


def add_customer(request):
    form = CustomerModelForm()

    if request.method == 'POST':
        form = CustomerModelForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, 'Customer successfully added!.')
            return redirect('customers')

    context = {'form': form}
    return render(request, 'customer/add-customer.html', context)


def update_customer(request, customer_id):
    customer = Customer.objects.get(id=customer_id)
    form = CustomerModelForm(instance=customer)
    if request.method == 'POST':
        form = CustomerModelForm(request.POST, request.FILES, instance=customer)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, 'Customer successfully updated.')
            return redirect('customers')

    context = {'form': form}
    return render(request, 'customer/update-customer.html', context)


def delete_customer(request, customer_id):
    customer = Customer.objects.get(id=customer_id)

    if request.method == 'POST':
        customer.delete()
        messages.add_message(request, messages.SUCCESS, 'Customer successfully deleted.')
        return redirect('customers')
    return render(request, 'customer/delete-customer.html', {'customer': customer})


def export_data(request):

    format = request.GET.get('format', 'csv')
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customers.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Full Name', 'Phone number', 'Email', 'Address', 'Image'])
        for customer in Customer.objects.all():
            writer.writerow([customer.id, customer.fullname, customer.email, customer.phone,
                             customer.address, customer.image])
    elif format == 'json':
        response = HttpResponse(content_type='application/json')
        data = list(Customer.objects.all().values('id', 'fullname', 'phone', 'email', 'address', 'image'))
        response.write(json.dumps(data, indent=4))
        response['Content-Disposition'] = 'attachment; filename="customers.json"'
    elif format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Customers'

        # Write header row
        header = ['ID', 'Full name', 'Phone number', 'Email', 'Address', 'Image']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        queryset = Customer.objects.all().values_list('id', 'fullname', 'phone', 'email', 'address', 'image')
        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
    else:
        response = HttpResponse(status=404)
        response.content = 'Bad request'

    return response
